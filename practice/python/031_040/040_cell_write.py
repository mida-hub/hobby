from datetime import datetime
from openpyxl import Workbook

wb = Workbook()
ws = wb.active

ws['A1'] = 'セル1'

b1 = ws['B1']
b1.value = 3.1

c1 = ws['C1']
c1.value = datetime.now()

ws.title = 'セル書き込みワークシート'
wb.save('output/cell_write.xlsx')
